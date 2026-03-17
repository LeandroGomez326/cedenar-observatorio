from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json
from django.views.decorators.csrf import csrf_exempt
from .models import Proyecto, Medicion, EstadoServidor, ConfiguracionAlerta, PreferenciaDashboard  # ← IMPORTADO
from .forms import ProyectoForm
from monitoreo.services.inversores.huawei import HuaweiInversor
from monitoreo.services.inversores.growatt import GrowattInversor
from .decorators import rate_limit
import json
from django.shortcuts import render
from .models import Proyecto, Medicion
from django.shortcuts import render, redirect
from .models import Proyecto
from .services.csv_processor import procesar_archivo
from django.db import IntegrityError
from .services.predictor import PredictorEnergia
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from .decorators import rate_limit, rate_limit_api
from .models import ConsentimientoDatos
from django.views.decorators.csrf import csrf_protect
from django.conf import settings

@login_required
def politica_privacidad(request):
    """Vista de política de privacidad"""
    
    # Verificar si ya aceptó la versión actual
    ya_acepto = ConsentimientoDatos.objects.filter(
        usuario=request.user,
        version_politica='1.0'
    ).exists()
    
    context = {
        'ya_acepto': ya_acepto,
        'version': '1.0',
        'fecha_actualizacion': '2024-03-04',
    }
    return render(request, 'monitoreo/politica_privacidad.html', context)

@login_required
@csrf_protect
def aceptar_privacidad(request):
    """Vista para aceptar la política de privacidad"""
    if request.method == 'POST':
        ConsentimientoDatos.objects.create(
            usuario=request.user,
            ip_origen=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            version_politica='1.0'
        )
        messages.success(request, '✅ Gracias por aceptar nuestra política de privacidad.')
        
        # Redirigir a la página que intentaba acceder
        next_url = request.POST.get('next', 'dashboard')
        return redirect(next_url)
    
    return redirect('politica_privacidad')

@login_required
def entrenar_modelo(request, proyecto_id):
    """Entrena el modelo de IA para un proyecto"""
    if not settings.IA_HABILITADA:
        return JsonResponse({
            'status': 'error',
            'message': '⚠️ IA deshabilitada en configuración'
        }, status=400)
    
    try:
        predictor = PredictorEnergia(proyecto_id)
        resultado = predictor.entrenar(dias_historial=365)
        
        if resultado:
            return JsonResponse({
                'status': 'success',
                'message': '✅ Modelo entrenado correctamente'
            })
        else:
            return JsonResponse({
                'status': 'error',
                'message': '❌ No hay suficientes datos (mínimo 100 registros)'
            }, status=400)
    
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
@login_required
def predecir(request, proyecto_id):
    """Obtiene predicciones para un proyecto"""
    if not settings.IA_HABILITADA:
        return JsonResponse({
            'status': 'error',
            'message': '⚠️ IA deshabilitada en configuración'
        }, status=400)
    
    dias = int(request.GET.get('dias', 7))
    
    try:
        predictor = PredictorEnergia(proyecto_id)
        predicciones = predictor.predecir_proximos_dias(dias)
        
        if 'error' in predicciones:
            return JsonResponse({
                'status': 'error',
                'message': predicciones['error']
            }, status=400)
        
        return JsonResponse({
            'status': 'success',
            'data': predicciones
        })
    
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
def verificar_modelos_ia(request):
    """Devuelve lista de proyectos que tienen modelo entrenado"""
    import os
    from django.conf import settings
    
    modelos_path = os.path.join(settings.BASE_DIR, 'modelos_ia')
    proyectos_con_modelo = []
    
    if os.path.exists(modelos_path):
        # Listar archivos de modelo
        for archivo in os.listdir(modelos_path):
            if archivo.startswith('modelo_gen_') and archivo.endswith('.pkl'):
                try:
                    proyecto_id = int(archivo.replace('modelo_gen_', '').replace('.pkl', ''))
                    # Verificar que el proyecto existe
                    if Proyecto.objects.filter(id=proyecto_id).exists():
                        proyectos_con_modelo.append(proyecto_id)
                except:
                    pass
    
    return JsonResponse({
        'proyectos_con_modelo': proyectos_con_modelo
    })

@login_required
def guardar_preferencias(request):
    """Guarda las preferencias del dashboard"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            preferencias, _ = PreferenciaDashboard.objects.get_or_create(
                usuario=request.user
            )
            
            preferencias.tema = data.get('tema', preferencias.tema)
            preferencias.layout = data.get('layout', preferencias.layout)
            preferencias.widgets_visibles = data.get('widgets_visibles', [])
            preferencias.orden_widgets = data.get('orden_widgets', [])
            preferencias.save()
            
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required
def cargar_preferencias(request):
    """Carga las preferencias del usuario"""
    preferencias, _ = PreferenciaDashboard.objects.get_or_create(
        usuario=request.user
    )
    
    return JsonResponse({
        'tema': preferencias.tema,
        'layout': preferencias.layout,
        'widgets_visibles': preferencias.widgets_visibles,
        'orden_widgets': preferencias.orden_widgets
    })

@login_required
def dashboard_personalizable(request):
    """Vista del dashboard personalizable"""
    return render(request, 'monitoreo/dashboard_personalizable.html')
# ============================================
# VISTA DE REGISTRO DE USUARIOS
# ============================================
@rate_limit(key='ip', rate='5/m') 
def registro(request):
    """
    Vista para registro de nuevos usuarios
    """
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'¡Bienvenido {user.username}! Tu cuenta ha sido creada exitosamente.')
            return redirect('dashboard')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = UserCreationForm()
    
    return render(request, 'monitoreo/registro.html', {'form': form})


# ============================================
# VISTAS DE DASHBOARD Y DETALLE
# ============================================
@login_required
def dashboard(request):
    proyectos = Proyecto.objects.all()
    proyectos_con_datos = []
    
    for p in proyectos:
        # Obtener todas las mediciones del proyecto
        mediciones = Medicion.objects.filter(proyecto=p).order_by('fecha_lectura')
        
        if mediciones.count() >= 2:
            primera = mediciones.first()
            ultima = mediciones.last()
            
            # Calcular consumo total (Energia Activa+)
            consumo_total = ultima.energia_activa_import - primera.energia_activa_import
            
            # Calcular generación total (Energia Activa-)
            generacion_total = ultima.energia_activa_export - primera.energia_activa_export
        else:
            consumo_total = 0
            generacion_total = 0
        
        # Última medición para mostrar fecha
        ultima_medicion = mediciones.last()
        
        proyecto_data = {
            'id': p.id,
            'nombre': p.nombre,
            'marca': p.marca or 'Sin marca',
            'codigo_medidor': p.codigo_medidor,
            'ubicacion': p.ubicacion or 'No especificada',
            'activo': p.activo,
            'estado': 'activo' if p.activo else 'inactivo',
            'generacion': round(max(0, generacion_total), 2),
            'consumo': round(max(0, consumo_total), 2),
            'ultima_lectura': ultima_medicion.fecha_lectura.strftime('%d/%m/%Y %H:%M') if ultima_medicion else '',
        }
        proyectos_con_datos.append(proyecto_data)
    
    total_proyectos = len(proyectos_con_datos)
    proyectos_activos = sum(1 for p in proyectos_con_datos if p['activo'])
    
    # DEBUG
    print("="*50)
    print("📊 DASHBOARD - Datos calculados:")
    for p in proyectos_con_datos:
        print(f"  • {p['nombre']}:")
        print(f"    - Consumo (Energia Activa+): {p['consumo']} kWh")
        print(f"    - Generación (Energia Activa-): {p['generacion']} kWh")
    print("="*50)
    
    return render(request, 'monitoreo/dashboard.html', {
        'proyectos': proyectos_con_datos,
        'total_proyectos': total_proyectos,
        'proyectos_activos': proyectos_activos,
    })
@login_required
def detalle_proyecto(request, proyecto_id):
    """Vista de detalle de un proyecto específico"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener última medición
    ultima_medicion = Medicion.objects.filter(proyecto=proyecto).order_by('-fecha_lectura').first()
    
    # Calcular totales (última - primera de TODAS las mediciones)
    mediciones = Medicion.objects.filter(
        proyecto=proyecto
    ).order_by('fecha_lectura')
    
    if mediciones.count() > 1:
        primera = mediciones.first()
        ultima = mediciones.last()
        
        consumo_total = ultima.energia_activa_import - primera.energia_activa_import
        generacion_total = ultima.energia_activa_export - primera.energia_activa_export
    else:
        consumo_total = 0
        generacion_total = 0
    
    context = {
        'proyecto': proyecto,
        'ultima_medicion': ultima_medicion,
        'consumo_total': round(max(0, consumo_total), 2),
        'generacion_total': round(max(0, generacion_total), 2),
        'balance': round(generacion_total - consumo_total, 2),
        'fecha_actual': timezone.now(),
    }
    return render(request, 'monitoreo/detalle_proyecto.html', context)

# ============================================
# APIs PARA GRÁFICAS
# ============================================

@rate_limit_api(key='user', rate='100/h')  
def api_lecturas(request, proyecto_id):
    """API para obtener datos de lecturas - acepta ID o código_medidor"""
    
    # Intentar buscar por ID primero
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except (ValueError, Proyecto.DoesNotExist):
        # Si falla, intentar buscar por código_medidor
        try:
            proyecto = Proyecto.objects.get(codigo_medidor=proyecto_id)
        except Proyecto.DoesNotExist:
            return JsonResponse({'error': 'Proyecto no encontrado'}, status=404)
    
    # Parámetros de la solicitud
    dias = int(request.GET.get('dias', 30))
    fecha_inicio = timezone.now() - timedelta(days=dias)
    
    # Obtener lecturas
    lecturas = Medicion.objects.filter(
        proyecto=proyecto
    ).order_by('fecha_lectura')
    
    # Si hay más de 100 lecturas, limitar a los últimos 'dias' días
    if lecturas.count() > 100:
        lecturas = lecturas.filter(fecha_lectura__gte=fecha_inicio)
    
    # Preparar datos para la gráfica
    fechas = []
    consumo = []
    generacion = []
    valores_import = []
    valores_export = []
    
    for i, lectura in enumerate(lecturas):
        fecha_str = lectura.fecha_lectura.strftime('%Y-%m-%d %H:%M')
        fechas.append(fecha_str)
        valores_import.append(lectura.energia_activa_import)
        valores_export.append(lectura.energia_activa_export)
        
        # Calcular consumo y generación (diferencia con lectura anterior)
        if i > 0:
            consumo_hora = lectura.energia_activa_import - lecturas[i-1].energia_activa_import
            generacion_hora = lectura.energia_activa_export - lecturas[i-1].energia_activa_export
        else:
            consumo_hora = 0
            generacion_hora = 0
        
        # Asegurar que no sean negativos
        consumo_hora = max(0, consumo_hora)
        generacion_hora = max(0, generacion_hora)
        
        consumo.append(round(consumo_hora, 2))
        generacion.append(round(generacion_hora, 2))
    
    return JsonResponse({
        'proyecto': proyecto.nombre,
        'fechas': fechas,
        'consumo': consumo,
        'generacion': generacion,
        'valores_import': valores_import,
        'valores_export': valores_export,
        'total_consumo': round(sum(consumo), 2),
        'total_generacion': round(sum(generacion), 2),
        'balance': round(sum(generacion) - sum(consumo), 2)
    })

@rate_limit_api(key='user', rate='50/h')  # ← AGREGAR ESTO
def api_resumen(request):
    """API con resumen de todos los proyectos"""
    proyectos = Proyecto.objects.all()
    data = []
    
    for proyecto in proyectos:
        ultima_medicion = Medicion.objects.filter(
            proyecto=proyecto
        ).order_by('-fecha_lectura').first()
        
        # Totales de TODAS las mediciones
        mediciones = Medicion.objects.filter(
            proyecto=proyecto
        ).order_by('fecha_lectura')
        
        if mediciones.count() > 1:
            primera = mediciones.first()
            ultima = mediciones.last()
            consumo = ultima.energia_activa_import - primera.energia_activa_import
            generacion = ultima.energia_activa_export - primera.energia_activa_export
        else:
            consumo = 0
            generacion = 0
        
        data.append({
            'id': proyecto.id,
            'nombre': proyecto.nombre,
            'codigo_medidor': proyecto.codigo_medidor,
            'marca': proyecto.marca,
            'ubicacion': proyecto.ubicacion,
            'activo': proyecto.activo,
            'ultima_lectura': ultima_medicion.fecha_lectura if ultima_medicion else None,
            'consumo_total': round(max(0, consumo), 2),
            'generacion_total': round(max(0, generacion), 2),
        })
    
    return JsonResponse({'proyectos': data})

# ============================================
# VISTAS DE EXPORTACIÓN
# ============================================

@login_required
def exportar_excel(request, proyecto_id):
    """Exporta las mediciones de un proyecto a Excel"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mediciones"
    
    # Encabezados
    headers = ['Fecha', 'Hora', 'Consumo (kWh)', 'Generación (kWh)', 
               'Energía Importada', 'Energía Exportada']
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Obtener datos
    mediciones = Medicion.objects.filter(
        proyecto=proyecto
    ).order_by('fecha_lectura')
    
    # Calcular consumos y escribir datos
    row_num = 2
    consumo_acumulado = 0
    generacion_acumulada = 0
    
    for i, m in enumerate(mediciones):
        if i > 0:
            consumo = m.energia_activa_import - mediciones[i-1].energia_activa_import
            generacion = m.energia_activa_export - mediciones[i-1].energia_activa_export
        else:
            consumo = 0
            generacion = 0
        
        consumo = max(0, consumo)
        generacion = max(0, generacion)
        
        consumo_acumulado += consumo
        generacion_acumulada += generacion
        
        # Escribir fila
        ws.cell(row=row_num, column=1, value=m.fecha_lectura.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=2, value=m.fecha_lectura.strftime('%H:%M'))
        ws.cell(row=row_num, column=3, value=round(consumo, 2))
        ws.cell(row=row_num, column=4, value=round(generacion, 2))
        ws.cell(row=row_num, column=5, value=m.energia_activa_import)
        ws.cell(row=row_num, column=6, value=m.energia_activa_export)
        
        row_num += 1
    
    # Agregar fila de totales
    total_row = row_num + 1
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=round(consumo_acumulado, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=round(generacion_acumulada, 2)).font = Font(bold=True)
    
    # Ajustar ancho de columnas
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{proyecto.nombre}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

# ============================================
# VISTAS DE INVERSORES
# ============================================

@staff_member_required
def estado_inversores(request):
    """Vista para administradores - ver estado de conexión (MODO SIMULACIÓN)"""
    proyectos = Proyecto.objects.all()
    estados = []
    
    for proyecto in proyectos:
        # Última medición real
        ultima_medicion = Medicion.objects.filter(
            proyecto=proyecto
        ).order_by('-fecha_lectura').first()
        
        estado = {
            'proyecto': proyecto,
            'conectado': True,  # En simulación siempre conectado
            'mensaje': 'Modo Simulación',
            'ultima_medicion': ultima_medicion,
            'simulado': True
        }
        
        # Verificar si tiene datos recientes
        if ultima_medicion:
            horas_sin_datos = (timezone.now() - ultima_medicion.fecha_lectura).total_seconds() / 3600
            if horas_sin_datos > 24:
                estado['mensaje'] = f'Sin datos por {horas_sin_datos:.0f}h'
        
        estados.append(estado)
    
    return render(request, 'monitoreo/estado_inversores.html', {
        'estados': estados,
        'modo_simulacion': True
    })

# ============================================
# VISTAS DE PROYECTOS (CRUD)
# ============================================

@login_required
def lista_proyectos(request):
    """Lista todos los proyectos"""
    proyectos = Proyecto.objects.all().order_by('-activo', 'nombre')
    
    # Calcular estadísticas
    total_proyectos = proyectos.count()
    proyectos_activos = proyectos.filter(activo=True).count()
    proyectos_inactivos = total_proyectos - proyectos_activos
    
    # Para depuración (lo verás en la consola de Django)
    print("="*50)
    print("📊 ESTADÍSTICAS DE PROYECTOS:")
    print(f"  Total: {total_proyectos}")
    print(f"  Activos: {proyectos_activos}")
    print(f"  Inactivos: {proyectos_inactivos}")
    for p in proyectos:
        print(f"  - {p.nombre}: {'✅ Activo' if p.activo else '❌ Inactivo'}")
    print("="*50)
    
    return render(request, 'monitoreo/lista_proyectos.html', {
        'proyectos': proyectos,
        'total_proyectos': total_proyectos,
        'proyectos_activos': proyectos_activos,
        'proyectos_inactivos': proyectos_inactivos,
    })
@rate_limit(key='user', rate='10/m')  
def crear_proyecto(request):
    """Crea un nuevo proyecto"""
    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            try:
                proyecto = form.save()
                messages.success(request, f'✅ Proyecto "{proyecto.nombre}" creado exitosamente.')
                return redirect('lista_proyectos')
            except IntegrityError:
                # Si hay error de duplicado, reintentar con un nuevo ID
                from django.db import connection
                with connection.cursor() as cursor:
                    if connection.vendor == 'postgresql':
                        cursor.execute("SELECT setval('monitoreo_proyecto_id_seq', (SELECT MAX(id) FROM monitoreo_proyecto), true);")
                    else:
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name='monitoreo_proyecto';")
                        cursor.execute("INSERT INTO sqlite_sequence (name, seq) VALUES ('monitoreo_proyecto', (SELECT MAX(id) FROM monitoreo_proyecto));")
                
                # Reintentar guardar
                proyecto = form.save()
                messages.success(request, f'✅ Proyecto "{proyecto.nombre}" creado exitosamente (secuencia corregida).')
                return redirect('lista_proyectos')
    else:
        form = ProyectoForm()
    
    return render(request, 'monitoreo/proyecto_form.html', {
        'form': form,
        'titulo': 'Crear Nuevo Proyecto'
    })
@login_required
def editar_proyecto(request, proyecto_id):
    """Edita un proyecto existente"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)  # ← 'proyecto' se define AQUÍ
    
    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=proyecto)
        if form.is_valid():
            proyecto = form.save()
            messages.success(request, f'Proyecto "{proyecto.nombre}" actualizado.')
            return redirect('lista_proyectos')
    else:
        form = ProyectoForm(instance=proyecto)
    
    # Aquí SÍ podemos usar 'proyecto' porque existe
    return render(request, 'monitoreo/proyecto_form.html', {
        'form': form,
        'proyecto': proyecto,  # ← Solo existe en editar
        'titulo': f'Editar Proyecto: {proyecto.nombre}'
    })

@login_required
def eliminar_proyecto(request, proyecto_id):
    """Elimina un proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        nombre = proyecto.nombre
        proyecto.delete()
        messages.success(request, f'Proyecto "{nombre}" eliminado.')
        return redirect('lista_proyectos')
    
    return render(request, 'monitoreo/proyecto_confirmar_eliminar.html', {
        'proyecto': proyecto
    })

@login_required
def toggle_proyecto_activo(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    proyecto.activo = not proyecto.activo
    proyecto.save()
    
    estado = "activado" if proyecto.activo else "desactivado"
    messages.info(request, f'Proyecto "{proyecto.nombre}" {estado}.')
    return redirect('lista_proyectos')  # ← Sin 'monitoreo:'

# ============================================
# VISTAS DE ESTADO DE SERVIDORES
# ============================================

@login_required
def estado_servidores(request):
    """Vista de estado de servidores"""
    estados = EstadoServidor.objects.all().order_by('nombre')
    
    # Si no hay registros, crear algunos por defecto
    if not estados.exists():
        for tipo, _ in EstadoServidor.TIPO_CHOICES:
            EstadoServidor.objects.get_or_create(nombre=tipo)
        estados = EstadoServidor.objects.all().order_by('nombre')
    
    # Estadísticas generales
    total_servicios = estados.count()
    servicios_activos = estados.filter(activo=True).count()
    porcentaje_uptime = (servicios_activos / total_servicios * 100) if total_servicios > 0 else 0
    
    return render(request, 'monitoreo/estado_servidores.html', {
        'estados': estados,
        'total_servicios': total_servicios,
        'servicios_activos': servicios_activos,
        'porcentaje_uptime': round(porcentaje_uptime, 1)
    })
@login_required
def campana_generacion(request, proyecto_id):
    """Vista de curva de generación diaria (forma de campana)"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    fecha_actual = timezone.now().strftime('%Y-%m-%d')
    
    return render(request, 'monitoreo/campana_generacion.html', {
        'proyecto': proyecto,
        'fecha_actual': fecha_actual
    })

@login_required
def api_generacion_diaria(request, proyecto_id):
    """API para obtener la curva de generación diaria promedio"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener fecha de la solicitud o usar hoy
    fecha = request.GET.get('fecha')
    if fecha:
        fecha_inicio = datetime.strptime(fecha, '%Y-%m-%d')
        fecha_fin = fecha_inicio + timedelta(days=1)
    else:
        fecha_inicio = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        fecha_fin = fecha_inicio + timedelta(days=1)
    
    # Hacer fecha_inicio aware si es naive
    if timezone.is_naive(fecha_inicio):
        fecha_inicio = timezone.make_aware(fecha_inicio)
    if timezone.is_naive(fecha_fin):
        fecha_fin = timezone.make_aware(fecha_fin)
    
    # Obtener mediciones del día
    mediciones = Medicion.objects.filter(
        proyecto=proyecto,
        fecha_lectura__gte=fecha_inicio,
        fecha_lectura__lt=fecha_fin
    ).order_by('fecha_lectura')
    
    # Preparar datos horarios
    horas = []
    generacion = []
    
    # Crear array de 24 horas con valores por defecto 0
    for h in range(24):
        horas.append(f"{h:02d}:00")
        generacion.append(0)
    
    # Calcular generación por hora (diferencia entre lecturas)
    for i in range(1, len(mediciones)):
        hora = mediciones[i].fecha_lectura.hour
        gen_hora = mediciones[i].energia_activa_export - mediciones[i-1].energia_activa_export
        generacion[hora] = round(max(0, gen_hora), 2)
    
    return JsonResponse({
        'proyecto': proyecto.nombre,
        'fecha': fecha_inicio.strftime('%Y-%m-%d'),
        'horas': horas,
        'generacion': generacion,
        'total_dia': round(sum(generacion), 2)
    })
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
from datetime import datetime, timedelta

@login_required
def reporte_pdf_mensual(request, proyecto_id):
    """Genera un reporte mensual en PDF"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener mes y año de los parámetros (o usar mes actual)
    mes = int(request.GET.get('mes', datetime.now().month))
    año = int(request.GET.get('año', datetime.now().year))
    
    # Calcular inicio y fin del mes
    inicio_mes = datetime(año, mes, 1)
    if mes == 12:
        fin_mes = datetime(año + 1, 1, 1) - timedelta(days=1)
    else:
        fin_mes = datetime(año, mes + 1, 1) - timedelta(days=1)
    
    # Hacer fechas aware
    inicio_mes = timezone.make_aware(inicio_mes)
    fin_mes = timezone.make_aware(fin_mes.replace(hour=23, minute=59, second=59))
    
    # Obtener mediciones del mes
    mediciones = Medicion.objects.filter(
        proyecto=proyecto,
        fecha_lectura__range=[inicio_mes, fin_mes]
    ).order_by('fecha_lectura')
    
    # Calcular totales
    if mediciones.count() > 1:
        primera = mediciones.first()
        ultima = mediciones.last()
        consumo_mes = ultima.energia_activa_import - primera.energia_activa_import
        generacion_mes = ultima.energia_activa_export - primera.energia_activa_export
    else:
        consumo_mes = 0
        generacion_mes = 0
    
    # Calcular datos diarios
    datos_diarios = []
    dia_actual = inicio_mes
    while dia_actual <= fin_mes:
        dia_siguiente = dia_actual + timedelta(days=1)
        
        lecturas_dia = Medicion.objects.filter(
            proyecto=proyecto,
            fecha_lectura__range=[dia_actual, dia_siguiente]
        ).order_by('fecha_lectura')
        
        if lecturas_dia.count() >= 2:
            primera_dia = lecturas_dia.first()
            ultima_dia = lecturas_dia.last()
            consumo_dia = ultima_dia.energia_activa_import - primera_dia.energia_activa_import
            generacion_dia = ultima_dia.energia_activa_export - primera_dia.energia_activa_export
        else:
            consumo_dia = 0
            generacion_dia = 0
        
        datos_diarios.append({
            'fecha': dia_actual.strftime('%d/%m/%Y'),
            'consumo': round(max(0, consumo_dia), 2),
            'generacion': round(max(0, generacion_dia), 2),
            'balance': round(max(0, generacion_dia) - max(0, consumo_dia), 2)
        })
        
        dia_actual = dia_siguiente
    
    # Calcular estadísticas
    consumos = [d['consumo'] for d in datos_diarios if d['consumo'] > 0]
    generaciones = [d['generacion'] for d in datos_diarios if d['generacion'] > 0]
    
    contexto = {
        'proyecto': proyecto,
        'mes': mes,
        'año': año,
        'nombre_mes': datetime(año, mes, 1).strftime('%B').capitalize(),
        'consumo_mes': round(consumo_mes, 2),
        'generacion_mes': round(generacion_mes, 2),
        'balance_mes': round(generacion_mes - consumo_mes, 2),
        'promedio_consumo_diario': round(sum(consumos) / len(consumos), 2) if consumos else 0,
        'promedio_generacion_diario': round(sum(generaciones) / len(generaciones), 2) if generaciones else 0,
        'max_consumo_dia': max(consumos) if consumos else 0,
        'max_generacion_dia': max(generaciones) if generaciones else 0,
        'dias_con_datos': len(consumos),
        'datos_diarios': datos_diarios,
        'generado_en': datetime.now().strftime('%d/%m/%Y %H:%M'),
    }
    
    # Renderizar HTML
    html_string = render_to_string('monitoreo/reporte_pdf.html', contexto)
    
    # Generar PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{proyecto.nombre}_{mes}_{año}.pdf"'
    
    HTML(string=html_string).write_pdf(response)
    
    return response

@login_required
def mapa_proyectos(request):
    """Vista del mapa de proyectos"""
    # Obtener TODOS los proyectos
    proyectos = Proyecto.objects.all()
    
    # Calcular totales para cada proyecto (opcional)
    for proyecto in proyectos:
        mediciones = Medicion.objects.filter(proyecto=proyecto).order_by('fecha_lectura')
        if mediciones.count() > 1:
            primera = mediciones.first()
            ultima = mediciones.last()
            proyecto.consumo_total = round(max(0, ultima.energia_activa_import - primera.energia_activa_import), 2)
            proyecto.generacion_total = round(max(0, ultima.energia_activa_export - primera.energia_activa_export), 2)
        else:
            proyecto.consumo_total = 0
            proyecto.generacion_total = 0
    
    print(f"📊 Vista mapa: {proyectos.count()} proyectos encontrados")  # DEBUG
    
    return render(request, 'monitoreo/mapa_proyectos.html', {
        'proyectos': proyectos
    })

def mapa_nuevo(request):
    """Vista nueva y simple para el mapa"""
    proyectos = Proyecto.objects.all()
    return render(request, 'monitoreo/mapa_nuevo.html', {'proyectos': proyectos})

@rate_limit(key='user', rate='20/m')  # ← AGREGAR ESTO
def subir_csv(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        
        # Procesar el archivo (CSV o Excel)
        resultado = procesar_archivo(archivo, proyecto)
        
        if resultado['exito']:
            messages.success(request, resultado['mensaje'])
        else:
            messages.error(request, resultado['mensaje'])
        
        return redirect('lista_proyectos')
    
    return render(request, 'monitoreo/subir_csv.html', {'proyecto': proyecto})

@login_required
def variables_electricas(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    return render(request, 'monitoreo/variables_electricas.html', {'proyecto': proyecto})

@login_required
def api_variables_electricas(request, proyecto_id):
    """API para obtener variables eléctricas detalladas"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    dias = int(request.GET.get('dias', 7))
    fecha_inicio = timezone.now() - timedelta(days=dias)
    
    mediciones = Medicion.objects.filter(
        proyecto=proyecto,
        fecha_lectura__gte=fecha_inicio
    ).order_by('fecha_lectura')
    
    data = {
        'fechas': [],
        'potencia_dc': [],
        'potencia_reactiva': [],
        'potencia_aparente': [],
        'corriente_ac_total': [],
        'corriente_dc': [],
        'voltaje_ac': [],
        'voltaje_dc': [],
        'corriente_ac': [],
        'voltaje_entre_fases': [],
        'factor_potencia': []
    }
    
    for m in mediciones:
        data['fechas'].append(m.fecha_lectura.strftime('%Y-%m-%d %H:%M'))
        data['potencia_dc'].append(m.potencia_dc_w or 0)
        data['potencia_reactiva'].append(m.potencia_reactiva_var or 0)
        data['potencia_aparente'].append(m.potencia_aparente_va or 0)
        data['corriente_ac_total'].append(m.corriente_ac_total_a or 0)
        data['corriente_dc'].append(m.corriente_dc_a or 0)
        data['voltaje_ac'].append(m.voltaje_ac_v or 0)
        data['voltaje_dc'].append(m.voltaje_dc_v or 0)
        data['corriente_ac'].append(m.corriente_ac_a or 0)
        data['voltaje_entre_fases'].append(m.voltaje_entre_fases_v or 0)
        data['factor_potencia'].append(m.factor_potencia_pct or 0)    
    return JsonResponse(data)


@login_required
def api_lecturas_por_periodo(request):
    """API que devuelve lecturas filtradas por rango de fechas"""
    
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    proyectos_ids = request.GET.getlist('proyectos[]')
    
    # Convertir fechas
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
    else:
        fecha_inicio = timezone.now() - timedelta(days=30)
        
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
    else:
        fecha_fin = timezone.now()
    
    # Hacer fechas aware
    fecha_inicio = timezone.make_aware(fecha_inicio)
    fecha_fin = timezone.make_aware(fecha_fin.replace(hour=23, minute=59, second=59))
    
    # Filtrar proyectos
    proyectos = Proyecto.objects.all()
    if proyectos_ids:
        proyectos = proyectos.filter(id__in=proyectos_ids)
    
    resultado = []
    
    for p in proyectos:
        mediciones = Medicion.objects.filter(
            proyecto=p,
            fecha_lectura__gte=fecha_inicio,
            fecha_lectura__lte=fecha_fin
        ).order_by('fecha_lectura')
        
        if mediciones.count() > 1:
            primera = mediciones.first()
            ultima = mediciones.last()
            consumo = ultima.energia_activa_import - primera.energia_activa_import
            generacion = ultima.energia_activa_export - primera.energia_activa_export
        else:
            consumo = 0
            generacion = 0
        
        resultado.append({
            'id': p.id,
            'nombre': p.nombre,
            'consumo_total': round(consumo, 2),
            'generacion_total': round(generacion, 2),
            'activo': p.activo
        })
    
    return JsonResponse({'proyectos': resultado})

import requests
from django.http import JsonResponse

def mi_ip(request):
    """
    Endpoint temporal para conocer la IP de salida de Render.
    """
    try:
        ip = requests.get('https://api.ipify.org').text
        return JsonResponse({
            'ip': ip,
            'mensaje': 'Agregá esta IP a la ACL de Oracle Cloud como XXX.XXX.XXX.XXX/32'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)